import pandas as p
import re
import openpyxl
from openpyxl.styles import Font as f
from openpyxl.utils import get_column_letter

df = p.read_excel("input.xlsx")
df.insert(0, "Title", range(1, len(df) + 1))
df['Fullname'] = df['ชื่อ - นามสกุล']
df['Address'] = df['ที่อยู่']

def SplitName(fullname):
    parts = fullname.strip().split()
    if len(parts) >= 3:
        return p.Series([parts[0], parts[1], " ".join(parts[2:])])
df[['Prefix', 'Name', 'Username']] = df['ชื่อ - นามสกุล'].apply(SplitName)

def SplitAddress(addr):
    addr = addr.replace("อตรอน", "อ.ตรอน").replace("อ.ป่าคาย","ต.ป่าคาย")
    result = {"No.": "","Moo": "","Soi": "","Road": "","Subdistrict": "","District": "","Province": "","Zip code": ""}
    No = re.search(r"\b(\d+\/?\d*)\b", addr)
    if No:
        result["No."] = No.group(1).strip()
    moo = re.search(r"(หมู่ที่|หมู่|ม\.)\s*(\d+)", addr)
    if moo:
        result["Moo"] = moo.group(2).strip()
    soi = re.search(r"(ซอย|ซ\.\s*)\s*([^\s]+(?:\s*\d*)?)", addr)
    if soi:
        result["Soi"] = soi.group(2).strip()
    road = re.search(r"(ถนน|ถ\.)\s*([^\s,]+)", addr)
    if road:
        result["Road"] = road.group(2).strip()
    tambon = re.search(r"(ต\.|ตำบล|แขวง)\s*([^\s,]+)", addr)
    if tambon:
        result["Subdistrict"] = tambon.group(2).strip()
    amphoe = re.search(r"(อ\.|อำเภอ|เขต)\s*([^\s,]+)", addr)
    if amphoe:
        result["District"] = amphoe.group(2).strip()
    province = re.search(r"(จ\.|จังหวัด)\s*([^\s,]+)", addr)
    if province:
        result["Province"] = province.group(2).strip()
    zip = re.search(r"\b\d{5}\b", addr)
    if zip:
        result["Zip code"] = zip.group(0).strip()
    return p.Series(result)

address = df["ที่อยู่"].apply(SplitAddress)

final = p.concat([df[['Title', 'Fullname', 'Address', 'Prefix', 'Name', 'Username']],address], axis=1)
# ได้ไฟล์ csv กับ xlsx
final.to_excel("master_table.xlsx", index=False)
final.to_csv("master_table.csv", index=False)

wb = openpyxl.load_workbook("master_table.xlsx")
ws = wb.active
font = f(name='Angsana New', size=13)
for col_idx, col in enumerate(ws.columns, 1):
    max_len = 0
    for cell in col:
        cell.font = font
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 2, 12)
wb.save("master_table.xlsx")

print("save master_table.xlsx & master_table.csv complete")
